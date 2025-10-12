#!/usr/bin/env python3
"""
Script to convert benchmark CSV results to Excel format with charts
"""

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_excel_report(csv_file="benchmark_results.csv", excel_file="benchmark_results.xlsx"):
    """
    Creates an Excel report from benchmark CSV data with charts and formatting
    """

    # Read CSV data
    if not os.path.exists(csv_file):
        print(f"Error: {csv_file} not found!")
        return False

    df = pd.read_csv(csv_file)

    # Create workbook and worksheets
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create raw data sheet
    ws_data = wb.create_sheet("Raw Data")

    # Add headers with formatting
    headers = ["Language", "Iteration", "Time(seconds)"]
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Add data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws_data.cell(row=row_idx, column=1, value=row['Language'])
        ws_data.cell(row=row_idx, column=2, value=row['Iteration'])
        ws_data.cell(row=row_idx, column=3, value=float(row['Time(seconds)']))

    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_data.column_dimensions[column_letter].width = adjusted_width

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")

    # Calculate averages for summary
    averages = df[df['Iteration'] == 'Average']
    individual_runs = df[df['Iteration'] != 'Average']

    # Summary headers
    summary_headers = ["Language", "Average Time (s)", "Min Time (s)", "Max Time (s)", "Std Dev (s)"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Calculate statistics for each language
    languages = ['Java', 'Python', 'JavaScript']
    for row_idx, lang in enumerate(languages, 2):
        lang_data = individual_runs[individual_runs['Language'] == lang]['Time(seconds)'].astype(float)

        ws_summary.cell(row=row_idx, column=1, value=lang)
        ws_summary.cell(row=row_idx, column=2, value=round(lang_data.mean(), 6))
        ws_summary.cell(row=row_idx, column=3, value=round(lang_data.min(), 6))
        ws_summary.cell(row=row_idx, column=4, value=round(lang_data.max(), 6))
        ws_summary.cell(row=row_idx, column=5, value=round(lang_data.std(), 6))

    # Auto-adjust column widths for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_summary.column_dimensions[column_letter].width = adjusted_width

    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Average Execution Time by Language"
    chart.y_axis.title = 'Time (seconds)'
    chart.x_axis.title = 'Programming Language'

    # Add data to chart
    data = Reference(ws_summary, min_col=2, min_row=1, max_row=4, max_col=2)
    cats = Reference(ws_summary, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Add chart to worksheet
    ws_summary.add_chart(chart, "G2")

    # Add title and metadata
    ws_summary.cell(row=6, column=1, value="Benchmark Report").font = Font(size=16, bold=True)
    ws_summary.cell(row=7, column=1, value=f"Total iterations per language: 10")
    ws_summary.cell(row=8, column=1, value=f"Generated from: {csv_file}")

    # Set active sheet to summary
    wb.active = ws_summary

    # Save workbook
    wb.save(excel_file)
    
    print(f"Excel report saved to: {excel_file}")

    # Create matplotlib charts
    create_matplotlib_charts(individual_runs)

    return True

def create_matplotlib_charts(df):
    """
    Create additional charts using matplotlib
    """
    plt.style.use('seaborn-v0_8')

    # Create figure with subplots
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 12))
    fig.suptitle('Benchmark Results Analysis', fontsize=16, fontweight='bold')

    # Chart 1: Average execution time by language (bar chart)
    avg_times = df.groupby('Language')['Time(seconds)'].mean()
    bars1 = ax1.bar(avg_times.index, avg_times.values, color=['#FF6B6B', '#4ECDC4', '#45B7D1'])
    ax1.set_title('Average Execution Time by Language')
    ax1.set_ylabel('Time (seconds)')
    ax1.set_xlabel('Programming Language')

    # Add value labels on bars
    for bar in bars1:
        height = bar.get_height()
        ax1.annotate(f'{height:.4f}s',
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),  # 3 points vertical offset
                    textcoords="offset points",
                    ha='center', va='bottom')

    # Chart 2: Box plot showing distribution
    languages = ['Java', 'Python', 'JavaScript']
    data_for_box = [df[df['Language'] == lang]['Time(seconds)'].values for lang in languages]
    bp = ax2.boxplot(data_for_box, labels=languages, patch_artist=True)
    ax2.set_title('Execution Time Distribution')
    ax2.set_ylabel('Time (seconds)')
    ax2.set_xlabel('Programming Language')

    # Color the box plots
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
    for patch, color in zip(bp['boxes'], colors):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)

    # Chart 3: Line plot showing performance across iterations
    for lang in languages:
        lang_data = df[df['Language'] == lang].sort_values('Iteration')
        ax3.plot(lang_data['Iteration'], lang_data['Time(seconds)'],
                marker='o', label=lang, linewidth=2, markersize=6)

    ax3.set_title('Performance Across Iterations')
    ax3.set_ylabel('Time (seconds)')
    ax3.set_xlabel('Iteration')
    ax3.legend()
    ax3.grid(True, alpha=0.3)

    # Chart 4: Performance statistics table
    ax4.axis('tight')
    ax4.axis('off')

    stats_data = []
    for lang in languages:
        lang_times = df[df['Language'] == lang]['Time(seconds)']
        stats_data.append([
            lang,
            f"{lang_times.mean():.6f}",
            f"{lang_times.min():.6f}",
            f"{lang_times.max():.6f}",
            f"{lang_times.std():.6f}"
        ])

    table = ax4.table(cellText=stats_data,
                     colLabels=['Language', 'Mean (s)', 'Min (s)', 'Max (s)', 'Std Dev (s)'],
                     cellLoc='center',
                     loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 1.5)

    # Style the table
    for i in range(len(languages)):
        table[(i+1, 0)].set_facecolor(colors[i])
        table[(i+1, 0)].set_text_props(weight='bold', color='white')

    plt.tight_layout()
    plt.savefig("benchmark_charts.png")
    plt.show()

if __name__ == "__main__":
    print("Converting benchmark results to Excel format...")

    # Check if required packages are available
    try:
        import openpyxl
        import matplotlib
        import seaborn
    except ImportError as e:
        print(f"Missing required package: {e}")
        print("Please install required packages:")
        print("pip install pandas openpyxl matplotlib seaborn")
        exit(1)

    success = create_excel_report()
    if success:
        print("\nExcel report generation completed successfully!")
    else:
        print("\nFailed to generate Excel report.")